import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
import os

# 評価スコアを除去する関数
def clean_text(text):
    if not isinstance(text, str):
        return ""
    return re.sub(r'^\d\.\d\(\d+\)\s*\·\s*', '', text).strip()

# ファイルを整形する関数
def process_file(uploaded_file):
    wb = load_workbook(filename=uploaded_file, data_only=True)
    ws = wb.active
    rows = [cell[0].value for cell in ws.iter_rows(min_col=1, max_col=1)]

    website_url_map = {
        cell[0].row: cell[0].hyperlink.target
        for cell in ws.iter_rows(min_col=1, max_col=1)
        if cell[0].value and 'ウェブサイト' in str(cell[0].value) and cell[0].hyperlink
    }

    start_indices = [
        i for i, row in enumerate(rows)
        if isinstance(row, str) and any(k in row for k in ['株式会社', '㈱', '工場', '有限会社'])
    ]
    start_indices.append(len(rows))

    address_keywords = ['丁目', '番地', '町', '区', '市', '村', '−', '-', '通り']

    company_entries = []
    for i in range(len(start_indices) - 1):
        start, end = start_indices[i], start_indices[i + 1]
        block = rows[start:end]

        name = block[0].strip() if block else ''
        raw_industry = next((b for b in block[1:] if '·' in str(b) or 'メーカー' in str(b) or '製造' in str(b)), '')
        industry = clean_text(str(raw_industry))

        raw_address = next((line.strip() for line in block[1:] if any(k in str(line) for k in address_keywords)), '')
        address = clean_text(str(raw_address))

        phone_match = next((re.search(r'\d{2,4}-\d{2,4}-\d{3,4}', str(b)) for b in block if re.search(r'\d{2,4}-\d{2,4}-\d{3,4}', str(b) or '')), None)
        phone = phone_match.group() if phone_match else ''

        website_url = ''
        for r in range(start + 1, end + 1):
            if r + 1 in website_url_map:
                website_url = website_url_map[r + 1]
                break

        company_entries.append({
            '企業名': name,
            '業種': industry,
            '住所': address,
            '電話番号': phone,
            'WebサイトURL': website_url
        })

    df = pd.DataFrame(company_entries)
    df = df.drop_duplicates(subset=['電話番号'], keep='first')

    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return output, df

# Streamlit UI
st.set_page_config(page_title="Google企業リスト自動整形ツール", layout="centered")
st.title("📊 Google企業リスト自動整形ツール")
st.caption("開発：株式会社トータルワーカーズ")

uploaded_file = st.file_uploader("📁 Excelファイルをアップロードしてください", type=["xlsx"])

if uploaded_file:
    with st.spinner("整形処理中..."):
        output_file, df = process_file(uploaded_file)
    st.success("整形が完了しました！")
    st.dataframe(df)

    st.download_button(
        label="📥 整形済みExcelをダウンロード",
        data=output_file,
        file_name="企業情報_cleaned.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )